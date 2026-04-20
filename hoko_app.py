import math
from io import BytesIO

import pandas as pd
import streamlit as st
import openpyxl

st.set_page_config(page_title="HOKO Mobility Financial Copilot v05", layout="wide")


# =========================================================
# FORMAT HELPERS
# =========================================================

def up_int(x):
    if pd.isna(x):
        return 0
    return int(math.ceil(float(x)))


def safe_float(x, default=0.0):
    try:
        if x is None or pd.isna(x):
            return default
        return float(x)
    except Exception:
        return default


def safe_int(x, default=0):
    try:
        if x is None or pd.isna(x):
            return default
        return int(math.ceil(float(x)))
    except Exception:
        return default


def fmt_num(x):
    return f"{up_int(x):,}"


def fmt_try(x):
    return f"₺{up_int(x):,}"


def fmt_pct(x):
    return f"%{up_int(x)}"


def fmt_plain_int(x):
    return str(up_int(x))


def is_yellow_fill(cell):
    if cell.fill is None or cell.fill.fill_type is None:
        return False
    rgb = cell.fill.fgColor.rgb
    if rgb is None:
        return False
    return str(rgb).upper() in ["FFFFFF00", "FFFF00"]


# =========================================================
# TEMPLATE EXCEL
# =========================================================

def make_template_excel():
    assumptions_rows = [
        ["", None, "Production Assumptions"],
        ["", 1, "Grace Period (Months)"],
        ["", 2, "Raw Material Required Before (Months)"],
        ["", 10, "Stock-In Rate %"],
        ["", None, "e-Scooter Assumptions"],
        ["", 200, "Battery Cost Per kWh (USD)"],
        ["", 3.24, "Battery kWh Each"],
        ["", 2, "Battery Per e-Scooter"],
        ["", 40.0, "Wh Consumption Per Km"],
        ["", 8, "Battery Life (Years)"],
        ["", None, "Charging Station Assumptions"],
        ["", 5000, "CS24 Hardware & Installation Cost (USD)"],
        ["", 25, "Franchisee Down Payment %"],
        ["", 0, "Franchisee Investment per CS24 (Calculated)"],
        ["", 3.5, "Energy Buy Price (TRY - 2026)"],
        ["", 15.0, "Energy Sell Price (TRY - 2026)"],
        ["", 0, "Gross Profit (TRY) (Calculated)"],
        ["", 60, "Franchesee Share %"],
        ["", 24, "Battery Slots Per CS24"],
        ["", 85, "Delivery SoC %"],
        ["", 2, "Charging Cycle Per CS24"],
        ["", 0.0, "ROI for Franchisee (Calculated)"],
        ["", 0.0, "IRR for Franchisee (Calculated)"],
        ["", None, "Operation Assumptions"],
        ["", 70, "Active User % (Courier)"],
        ["", 40, "Heavy Duty User % (Courier)"],
        ["", 180, "Heavy Duty km Per Day (Courier)"],
        ["", 40, "Standard Duty User % (Courier)"],
        ["", 140, "Standart Duty km Per Day (Courier)"],
        ["", 20, "Light Duty User % (Courier)"],
        ["", 100, "Light Duty km Per Day (Courier)"],
        ["", 30, "Active User % (Commuter)"],
        ["", 35, "Commuter km Per Day"],
    ]
    assumptions = pd.DataFrame(assumptions_rows, columns=["A", "B", "C"])

    monthly_inputs = pd.DataFrame({
        "Month": list(range(1, 49)),
        "Courier E-Scooter (Facelift) Units": [0] * 48,
        "Courier E-Scooter (New) Units": [0] * 48,
        "Commuter E-Scooter Units": [0] * 48,
        "Active User % (Courier)": [70] * 48,
        "Active User % (Commuter)": [30] * 48,
    })

    macro = pd.DataFrame({
        "Year": [2026, 2027, 2028, 2029],
        "Average USD/TRY (H1)": [45, 52, 60, 68],
        "Average USD/TRY (H2)": [48, 56, 64, 72],
        "Average Inflation Rate (H1)": [20, 16, 12, 10],
        "Average Inflation Rate (H2)": [18, 14, 10, 8],
    })

    opex = pd.DataFrame({
        "Cost Item": ["Rent", "Marketing", "Software & IT", "HQ Expenses"],
        "Monthly Cost (2026)": [250000, 150000, 80000, 200000],
        "Type": ["Fixed", "Semi-Variable", "Fixed", "Fixed"],
        "Category": ["Admin", "Growth", "Tech", "Admin"],
        "Escalation Type": ["Inflation", "Inflation", "Inflation", "Inflation"],
        "Escalation Rate %": [0, 0, 0, 0],
        "Notes": ["", "", "", ""],
    })

    personnel_base = pd.DataFrame({
        "Role": [
            "Plant Manager (Fabrika Müdürü)",
            "Assembly Operator (Montaj Operatörü)",
            "Line Leader (Hat Lideri)",
            "Quality Control Operator (Kalite Kontrol Operatörü)",
            "Test Operator (Test Operatörü)",
            "Line Feeder / Internal Logistics Operator (Hat Besleme / İç Lojistik Operatörü)",
            "CEO",
            "Business Development Manager",
        ],
        "Base Salary (2026)": [120000, 35000, 45000, 40000, 38000, 36000, 450000, 180000],
        "Notes": ["", "", "", "", "", "", "", ""],
    })

    personnel_rules = pd.DataFrame({
        "Role": [
            "Plant Manager (Fabrika Müdürü)",
            "Assembly Operator (Montaj Operatörü)",
            "Line Leader (Hat Lideri)",
            "Quality Control Operator (Kalite Kontrol Operatörü)",
            "Test Operator (Test Operatörü)",
            "Line Feeder / Internal Logistics Operator (Hat Besleme / İç Lojistik Operatörü)",
            "CEO",
            "Business Development Manager",
        ],
        "First Hire Month (Incl. Grace)": [1, 1, 1, 1, 1, 1, 6, 6],
        "Monthly (Units per Employee)": [1, 1000, 1000, 1000, 1000, 1000, 1, 3000],
        "Max Cap": [5, 500, 50, 50, 50, 50, 1, 10],
        "Base Headcount": [1, 1, 1, 1, 1, 1, 1, 1],
    })

    bom = pd.DataFrame({
        "Item": ["Frame", "Motor", "Controller", "Plastic Parts"],
        "Product": [
            "Courier E-Scooter (Facelift)",
            "Courier E-Scooter (New)",
            "Commuter E-Scooter",
            "Courier E-Scooter (New)"
        ],
        "Unit Cost USD": [120, 180, 90, 60],
    })

    product_capex = pd.DataFrame({
        "Product": [
            "Courier E-Scooter (Facelift)",
            "Courier E-Scooter (New)",
            "Commuter E-Scooter",
        ],
        "CapEx Per Unit USD": [700, 900, 650],
        "Battery kWh": [6.48, 6.48, 4.86],
        "Notes": ["", "", ""],
    })

    bio = BytesIO()
    with pd.ExcelWriter(bio, engine="openpyxl") as writer:
        assumptions.to_excel(writer, sheet_name="Assumptions", index=False, header=False)
        monthly_inputs.to_excel(writer, sheet_name="Monthly Inputs", index=False)
        macro.to_excel(writer, sheet_name="Macro Assumptions", index=False)
        opex.to_excel(writer, sheet_name="OPEX", index=False)
        personnel_base.to_excel(writer, sheet_name="Personnel Base", index=False)
        personnel_rules.to_excel(writer, sheet_name="Personnel Monthly Rules", index=False)
        bom.to_excel(writer, sheet_name="BoM List", index=False)
        product_capex.to_excel(writer, sheet_name="Product CapEx", index=False)

    bio.seek(0)
    return bio.getvalue()


# =========================================================
# ASSUMPTION LAYOUT READER
# =========================================================

def read_assumptions_layout_from_workbook(file_obj):
    file_obj.seek(0)
    wb = openpyxl.load_workbook(file_obj, data_only=False)
    ws = wb["Assumptions"]

    rows = []
    for r in range(1, ws.max_row + 1):
        label = ws[f"C{r}"].value
        value = ws[f"B{r}"].value
        if label is None:
            continue

        label_str = str(label).strip()

        # section header
        if value is None and label_str:
            rows.append({
                "row_no": r,
                "kind": "section",
                "label": label_str,
                "raw_value": None,
                "decimal": False,
            })
            continue

        # normal assumption line
        decimal_flag = is_yellow_fill(ws[f"B{r}"])
        rows.append({
            "row_no": r,
            "kind": "line",
            "label": label_str,
            "raw_value": value,
            "decimal": decimal_flag,
        })

    return rows


def default_assumptions_layout():
    return [
        {"row_no": 2, "kind": "section", "label": "Production Assumptions", "raw_value": None, "decimal": False},
        {"row_no": 3, "kind": "line", "label": "Grace Period (Months)", "raw_value": 1, "decimal": False},
        {"row_no": 4, "kind": "line", "label": "Raw Material Required Before (Months)", "raw_value": 2, "decimal": False},
        {"row_no": 5, "kind": "line", "label": "Stock-In Rate %", "raw_value": 10, "decimal": False},

        {"row_no": 7, "kind": "section", "label": "e-Scooter Assumptions", "raw_value": None, "decimal": False},
        {"row_no": 8, "kind": "line", "label": "Battery Cost Per kWh (USD)", "raw_value": 200, "decimal": False},
        {"row_no": 9, "kind": "line", "label": "Battery kWh Each", "raw_value": 3.24, "decimal": True},
        {"row_no": 10, "kind": "line", "label": "Battery Per e-Scooter", "raw_value": 2, "decimal": False},
        {"row_no": 11, "kind": "line", "label": "Wh Consumption Per Km", "raw_value": 40.0, "decimal": True},
        {"row_no": 12, "kind": "line", "label": "Battery Life (Years)", "raw_value": 8, "decimal": False},

        {"row_no": 14, "kind": "section", "label": "Charging Station Assumptions", "raw_value": None, "decimal": False},
        {"row_no": 15, "kind": "line", "label": "CS24 Hardware & Installation Cost (USD)", "raw_value": 5000, "decimal": False},
        {"row_no": 16, "kind": "line", "label": "Franchisee Down Payment %", "raw_value": 25, "decimal": False},
        {"row_no": 17, "kind": "line", "label": "Franchisee Investment per CS24 (Calculated)", "raw_value": 0, "decimal": False},
        {"row_no": 18, "kind": "line", "label": "Energy Buy Price (TRY - 2026)", "raw_value": 3.5, "decimal": True},
        {"row_no": 19, "kind": "line", "label": "Energy Sell Price (TRY - 2026)", "raw_value": 15.0, "decimal": True},
        {"row_no": 20, "kind": "line", "label": "Gross Profit (TRY) (Calculated)", "raw_value": 0.0, "decimal": True},
        {"row_no": 21, "kind": "line", "label": "Franchesee Share %", "raw_value": 60, "decimal": False},
        {"row_no": 22, "kind": "line", "label": "Battery Slots Per CS24", "raw_value": 24, "decimal": False},
        {"row_no": 23, "kind": "line", "label": "Delivery SoC %", "raw_value": 85, "decimal": False},
        {"row_no": 24, "kind": "line", "label": "Charging Cycle Per CS24", "raw_value": 2, "decimal": False},
        {"row_no": 25, "kind": "line", "label": "ROI for Franchisee (Calculated)", "raw_value": 0.0, "decimal": True},
        {"row_no": 26, "kind": "line", "label": "IRR for Franchisee (Calculated)", "raw_value": 0.0, "decimal": True},

        {"row_no": 28, "kind": "section", "label": "Operation Assumptions", "raw_value": None, "decimal": False},
        {"row_no": 29, "kind": "line", "label": "Active User % (Courier)", "raw_value": 70, "decimal": False},
        {"row_no": 30, "kind": "line", "label": "Heavy Duty User % (Courier)", "raw_value": 40, "decimal": False},
        {"row_no": 31, "kind": "line", "label": "Heavy Duty km Per Day (Courier)", "raw_value": 180, "decimal": False},
        {"row_no": 32, "kind": "line", "label": "Standard Duty User % (Courier)", "raw_value": 40, "decimal": False},
        {"row_no": 33, "kind": "line", "label": "Standart Duty km Per Day (Courier)", "raw_value": 140, "decimal": False},
        {"row_no": 34, "kind": "line", "label": "Light Duty User % (Courier)", "raw_value": 20, "decimal": False},
        {"row_no": 35, "kind": "line", "label": "Light Duty km Per Day (Courier)", "raw_value": 100, "decimal": False},
        {"row_no": 36, "kind": "line", "label": "Active User % (Commuter)", "raw_value": 30, "decimal": False},
        {"row_no": 37, "kind": "line", "label": "Commuter km Per Day", "raw_value": 35, "decimal": False},
    ]


# =========================================================
# ASSUMPTION VALUE HELPERS
# =========================================================

CALCULATED_LABELS = {
    "Franchisee Investment per CS24 (Calculated)",
    "Gross Profit (TRY) (Calculated)",
    "ROI for Franchisee (Calculated)",
    "IRR for Franchisee (Calculated)",
}

OVERRIDE_LABELS = {
    "Active User % (Courier)",
    "Heavy Duty User % (Courier)",
    "Heavy Duty km Per Day (Courier)",
    "Standard Duty User % (Courier)",
    "Standart Duty km Per Day (Courier)",
    "Light Duty User % (Courier)",
    "Light Duty km Per Day (Courier)",
    "Active User % (Commuter)",
    "Commuter km Per Day",
}


# =========================================================
# HEADER
# =========================================================

st.title("HOKO Mobility Financial Copilot v05")
st.caption("Excel-Driven, Multi-Product, Monthly Financial Model")

with st.expander("Excel Upload Instructions", expanded=True):
    st.markdown("""
**Please upload your Business Plan Excel file.**

Required sheets:
Assumptions  
Monthly Inputs  
Macro Assumptions  
OPEX  
Personnel Base  
Personnel Monthly Rules  
BoM List  
Product CapEx  

Cash Flow Statement and Balance Sheet are generated inside the app.
""")

st.download_button(
    "Download Template Excel",
    data=make_template_excel(),
    file_name="HOKO_BP_Template_v05.xlsx",
    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
)

uploaded_file = st.file_uploader("Upload Business Plan Excel", type=["xlsx"])


# =========================================================
# READ EXCEL
# =========================================================

assumptions_layout = default_assumptions_layout()
monthly_inputs_df = None
macro_df = None
opex_df = None
personnel_base_df = None
personnel_rules_df = None
bom_df = None
product_capex_df = None

if uploaded_file is not None:
    try:
        uploaded_file.seek(0)
        xls = pd.ExcelFile(uploaded_file)
        required = [
            "Assumptions",
            "Monthly Inputs",
            "Macro Assumptions",
            "OPEX",
            "Personnel Base",
            "Personnel Monthly Rules",
            "BoM List",
            "Product CapEx",
        ]
        missing = [s for s in required if s not in xls.sheet_names]
        if missing:
            st.error(f"Missing sheet(s): {', '.join(missing)}")
        else:
            assumptions_layout = read_assumptions_layout_from_workbook(uploaded_file)
            uploaded_file.seek(0)
            xls = pd.ExcelFile(uploaded_file)
            monthly_inputs_df = pd.read_excel(xls, "Monthly Inputs")
            macro_df = pd.read_excel(xls, "Macro Assumptions")
            opex_df = pd.read_excel(xls, "OPEX")
            personnel_base_df = pd.read_excel(xls, "Personnel Base")
            personnel_rules_df = pd.read_excel(xls, "Personnel Monthly Rules")
            bom_df = pd.read_excel(xls, "BoM List")
            product_capex_df = pd.read_excel(xls, "Product CapEx")
            st.success("Excel uploaded successfully.")
    except Exception as e:
        st.error(f"Excel could not be read: {e}")


# =========================================================
# ASSUMPTIONS UI (ORDER FROM EXCEL)
# =========================================================

st.markdown("## Assumptions")

assumption_inputs = {}

for row in assumptions_layout:
    if row["kind"] == "section":
        st.markdown(f"**{row['label']}**")
        continue

    label = row["label"]
    raw_value = row["raw_value"]
    decimal_flag = row["decimal"]

    if label in CALCULATED_LABELS:
        assumption_inputs[label] = {"type": "calculated", "value": raw_value, "decimal": decimal_flag}
        st.write(f"{label}: [Calculated]")
        continue

    if label in OVERRIDE_LABELS:
        default_val = "" if raw_value is None else str(raw_value)
        value = st.text_input(label, value=default_val, key=f"ovr_{label}")
        assumption_inputs[label] = {"type": "override", "value": value, "decimal": decimal_flag}
    else:
        if decimal_flag:
            value = st.number_input(
                label,
                value=safe_float(raw_value, 0.0),
                step=0.1,
                format="%.1f",
                key=f"ed_{label}",
            )
        else:
            value = st.number_input(
                label,
                value=safe_int(raw_value, 0),
                step=1,
                key=f"ed_{label}",
            )
        assumption_inputs[label] = {"type": "editable", "value": value, "decimal": decimal_flag}


# =========================================================
# ASSUMPTION EXTRACTION
# =========================================================

def get_assumption(label, default=0):
    if label not in assumption_inputs:
        return default
    item = assumption_inputs[label]
    if item["type"] == "override":
        return safe_float(item["value"], default) if item["decimal"] else safe_int(item["value"], default)
    return item["value"]


grace_period_months = safe_int(get_assumption("Grace Period (Months)", 1))
raw_material_required_before_months = safe_int(get_assumption("Raw Material Required Before (Months)", 2))
stock_in_rate_pct = safe_float(get_assumption("Stock-In Rate %", 10))

battery_cost_per_kwh_usd = safe_float(get_assumption("Battery Cost Per kWh (USD)", 200))
battery_kwh_each = safe_float(get_assumption("Battery kWh Each", 3.24))
battery_per_escooter = safe_int(get_assumption("Battery Per e-Scooter", 2))
wh_consumption_per_km = safe_float(get_assumption("Wh Consumption Per Km", 40.0))
battery_life_years = safe_int(get_assumption("Battery Life (Years)", 8))

cs24_hardware_installation_cost_usd = safe_float(get_assumption("CS24 Hardware & Installation Cost (USD)", 5000))
franchisee_down_payment_pct = safe_float(get_assumption("Franchisee Down Payment %", 25))
energy_buy_price_try_2026 = safe_float(get_assumption("Energy Buy Price (TRY - 2026)", 3.5))
energy_sell_price_try_2026 = safe_float(get_assumption("Energy Sell Price (TRY - 2026)", 15.0))
franchisee_share_pct = safe_float(get_assumption("Franchesee Share %", 60))
battery_slots_per_cs24 = safe_int(get_assumption("Battery Slots Per CS24", 24))
delivery_soc_pct = safe_float(get_assumption("Delivery SoC %", 85))
charging_cycle_per_cs24 = safe_float(get_assumption("Charging Cycle Per CS24", 2))

active_user_pct_courier = safe_float(get_assumption("Active User % (Courier)", 70))
heavy_duty_user_pct_courier = safe_float(get_assumption("Heavy Duty User % (Courier)", 40))
heavy_duty_km_per_day_courier = safe_float(get_assumption("Heavy Duty km Per Day (Courier)", 180))
standard_duty_user_pct_courier = safe_float(get_assumption("Standard Duty User % (Courier)", 40))
standard_duty_km_per_day_courier = safe_float(get_assumption("Standart Duty km Per Day (Courier)", 140))
light_duty_user_pct_courier = safe_float(get_assumption("Light Duty User % (Courier)", 20))
light_duty_km_per_day_courier = safe_float(get_assumption("Light Duty km Per Day (Courier)", 100))
active_user_pct_commuter = safe_float(get_assumption("Active User % (Commuter)", 30))
commuter_km_per_day = safe_float(get_assumption("Commuter km Per Day", 35))

if round(heavy_duty_user_pct_courier + standard_duty_user_pct_courier + light_duty_user_pct_courier, 4) != 100:
    st.warning(
        f"Heavy Duty + Standard Duty + Light Duty must equal %100. Current total: "
        f"%{math.ceil(heavy_duty_user_pct_courier + standard_duty_user_pct_courier + light_duty_user_pct_courier)}"
    )


# =========================================================
# CALCULATED ASSUMPTIONS
# =========================================================

weighted_avg_km_per_day_courier = (
    heavy_duty_km_per_day_courier * heavy_duty_user_pct_courier / 100
    + standard_duty_km_per_day_courier * standard_duty_user_pct_courier / 100
    + light_duty_km_per_day_courier * light_duty_user_pct_courier / 100
)

kwh_per_escooter_per_day = (
    weighted_avg_km_per_day_courier * active_user_pct_courier / 100
    + commuter_km_per_day * active_user_pct_commuter / 100
) * wh_consumption_per_km / 1000

franchisee_investment_per_cs24_calculated = (
    cs24_hardware_installation_cost_usd
    + (battery_slots_per_cs24 * battery_kwh_each * battery_cost_per_kwh_usd)
)

gross_profit_try_calculated = energy_sell_price_try_2026 - energy_buy_price_try_2026
roi_for_franchisee_calculated = 0.0
irr_for_franchisee_calculated = 0.0

calculated_map = {
    "Franchisee Investment per CS24 (Calculated)": franchisee_investment_per_cs24_calculated,
    "Gross Profit (TRY) (Calculated)": gross_profit_try_calculated,
    "ROI for Franchisee (Calculated)": roi_for_franchisee_calculated,
    "IRR for Franchisee (Calculated)": irr_for_franchisee_calculated,
}

# update assumptions display source
for row in assumptions_layout:
    if row["kind"] == "line" and row["label"] in calculated_map:
        row["raw_value"] = calculated_map[row["label"]]


# =========================================================
# PERSONNEL MAPS
# =========================================================

base_salary_map = {}
if personnel_base_df is not None and not personnel_base_df.empty:
    salary_col = None
    for c in personnel_base_df.columns:
        if "Base Salary" in str(c):
            salary_col = c
            break
    if salary_col:
        for _, row in personnel_base_df.iterrows():
            base_salary_map[str(row["Role"]).strip()] = safe_float(row[salary_col], 0)

personnel_rule_rows = []
if personnel_rules_df is not None and not personnel_rules_df.empty:
    for _, row in personnel_rules_df.iterrows():
        personnel_rule_rows.append({
            "Role": str(row["Role"]).strip(),
            "First Hire Month (Incl. Grace)": safe_int(row["First Hire Month (Incl. Grace)"], 1),
            "Monthly (Units per Employee)": max(1, safe_int(row["Monthly (Units per Employee)"], 1)),
            "Max Cap": max(1, safe_int(row["Max Cap"], 1)),
            "Base Headcount": max(1, safe_int(row["Base Headcount"], 1)),
        })


# =========================================================
# PRODUCT / BOM MAPS
# =========================================================

bom_monthly_usd_per_product = {}
if bom_df is not None and not bom_df.empty:
    grouped = bom_df.groupby("Product", as_index=False)["Unit Cost USD"].sum()
    for _, row in grouped.iterrows():
        bom_monthly_usd_per_product[str(row["Product"]).strip()] = safe_float(row["Unit Cost USD"], 0)


# =========================================================
# FALLBACK MONTHLY INPUTS
# =========================================================

def build_default_monthly_inputs():
    rows = []
    for m in range(1, 49):
        if 1 <= m <= 5:
            total = 0
        elif m == 6:
            total = 2000
        elif m == 7:
            total = 2500
        elif m in [8, 9]:
            total = 3000
        elif m in [10, 11, 12]:
            total = 3250
        elif 13 <= m <= 18:
            total = 3500
        else:
            total = 4000

        rows.append({
            "Month": m,
            "Courier E-Scooter (Facelift) Units": 0,
            "Courier E-Scooter (New) Units": total,
            "Commuter E-Scooter Units": 0,
            "Active User % (Courier)": 70,
            "Active User % (Commuter)": 30,
        })
    return pd.DataFrame(rows)


if monthly_inputs_df is None or monthly_inputs_df.empty:
    monthly_inputs_df = build_default_monthly_inputs()


# =========================================================
# MONTHLY ENGINE
# =========================================================

results = []
inventory_units = 0
inventory_raw_material_value = 0
active_escooters = 0
installed_cs24 = 0
cumulative_cash = 0
personnel_records = []
raw_material_purchase_schedule = {m: 0 for m in range(1, 49)}

for _, row in monthly_inputs_df.iterrows():
    month_no = safe_int(row["Month"], 1)
    year, half = get_half_for_month(month_no)

    usd_try_month = get_macro_value(macro_df, year, half, "Average USD/TRY")
    inflation_rate_month = get_macro_value(macro_df, year, half, "Average Inflation Rate")

    if usd_try_month == 0:
        usd_try_month = 45
    if inflation_rate_month == 0:
        inflation_rate_month = 15

    courier_facelift_units = safe_int(row.get("Courier E-Scooter (Facelift) Units", 0), 0)
    courier_new_units = safe_int(row.get("Courier E-Scooter (New) Units", 0), 0)
    commuter_units = safe_int(row.get("Commuter E-Scooter Units", 0), 0)

    monthly_production_units = courier_facelift_units + courier_new_units + commuter_units

    units_added_to_stock = math.ceil(monthly_production_units * stock_in_rate_pct / 100)
    units_sold_month = monthly_production_units - units_added_to_stock

    inventory_units += units_added_to_stock
    active_escooters += units_sold_month

    energy_buy_price_month = energy_buy_price_try_2026 * (1 + inflation_rate_month / 100)
    energy_sell_price_month = energy_sell_price_try_2026 * (1 + inflation_rate_month / 100)
    company_share_pct = 100 - franchisee_share_pct
    newco_margin_per_kwh_month = (energy_sell_price_month - energy_buy_price_month) * company_share_pct / 100

    monthly_kwh_sold = active_escooters * kwh_per_escooter_per_day * 30
    monthly_revenue = monthly_kwh_sold * energy_sell_price_month
    monthly_energy_cost = monthly_kwh_sold * energy_buy_price_month
    monthly_energy_gp_newco = monthly_kwh_sold * newco_margin_per_kwh_month

    monthly_motor_gross_profit = units_sold_month * 80000

    total_battery_kwh_per_cs24 = battery_slots_per_cs24 * battery_kwh_each
    sellable_kwh_per_cycle = total_battery_kwh_per_cs24 * delivery_soc_pct / 100
    daily_kwh_per_cs24 = sellable_kwh_per_cycle * charging_cycle_per_cs24

    monthly_cs24_required = math.ceil((active_escooters * kwh_per_escooter_per_day) / max(1, daily_kwh_per_cs24))
    monthly_new_cs24 = max(0, monthly_cs24_required - installed_cs24)
    installed_cs24 = monthly_cs24_required

    cs24_total_capex_try_month = (
        cs24_hardware_installation_cost_usd
        + battery_slots_per_cs24 * battery_kwh_each * battery_cost_per_kwh_usd
    ) * usd_try_month
    financed_per_cs24_try_month = cs24_total_capex_try_month * (1 - franchisee_down_payment_pct / 100)
    monthly_cs24_capex_outflow = monthly_new_cs24 * financed_per_cs24_try_month

    monthly_cs24_other_gp = installed_cs24 * ((1200 * usd_try_month) / 12)

    bom_usd_total = (
        courier_facelift_units * bom_monthly_usd_per_product.get("Courier E-Scooter (Facelift)", 0)
        + courier_new_units * bom_monthly_usd_per_product.get("Courier E-Scooter (New)", 0)
        + commuter_units * bom_monthly_usd_per_product.get("Commuter E-Scooter", 0)
    )
    current_raw_material_need_try = bom_usd_total * usd_try_month

    slowdown_factor = max(0, 1 - stock_in_rate_pct / 100)
    adjusted_raw_material_need_try = current_raw_material_need_try * slowdown_factor

    purchase_month = month_no - raw_material_required_before_months
    if 1 <= purchase_month <= 48:
        raw_material_purchase_schedule[purchase_month] += adjusted_raw_material_need_try

    raw_material_cost_month = raw_material_purchase_schedule.get(month_no, 0)
    inventory_raw_material_value = max(0, inventory_raw_material_value + raw_material_cost_month - adjusted_raw_material_need_try)

    total_personnel_cost_month = 0

    for pr in personnel_rule_rows:
        role = pr["Role"]
        actual_hire_month = pr["First Hire Month (Incl. Grace)"]

        if month_no < actual_hire_month:
            headcount = 0
        else:
            scaled = math.ceil(monthly_production_units / max(1, pr["Monthly (Units per Employee)"]))
            headcount = max(pr["Base Headcount"], scaled)
            headcount = min(headcount, pr["Max Cap"])

        salary_base = base_salary_map.get(role, 0)
        monthly_salary_actual = salary_base * (1 + inflation_rate_month / 100)
        monthly_personnel_cost = headcount * monthly_salary_actual

        total_personnel_cost_month += monthly_personnel_cost

        personnel_records.append({
            "Month": month_no,
            "Role": role,
            "Headcount": headcount,
            "Monthly Salary (Actual)": monthly_salary_actual,
            "Monthly Personnel Cost": monthly_personnel_cost,
        })

    total_opex_month = 0
    if opex_df is not None and not opex_df.empty:
        for _, orow in opex_df.iterrows():
            base_cost = safe_float(orow.get("Monthly Cost (2026)", 0), 0)
            esc_type = str(orow.get("Escalation Type", "")).strip().lower()
            custom_rate = safe_float(orow.get("Escalation Rate %", 0), 0)

            if esc_type == "custom":
                actual_cost = base_cost * (1 + custom_rate / 100)
            elif esc_type == "fx linked":
                actual_cost = base_cost * (usd_try_month / 45)
            else:
                actual_cost = base_cost * (1 + inflation_rate_month / 100)

            total_opex_month += actual_cost

    monthly_total_gross_profit = (
        monthly_energy_gp_newco + monthly_motor_gross_profit + monthly_cs24_other_gp
    )
    monthly_ebitda_proxy = (
        monthly_total_gross_profit - total_personnel_cost_month - total_opex_month
    )

    monthly_interest_cost = 0
    monthly_tax = max(0, monthly_ebitda_proxy) * 0.20
    monthly_net_income = monthly_ebitda_proxy - monthly_interest_cost - monthly_tax

    monthly_net_cash_flow = (
        monthly_net_income - monthly_cs24_capex_outflow - raw_material_cost_month
    )
    cumulative_cash += monthly_net_cash_flow

    installed_cs24_asset_value = installed_cs24 * cs24_total_capex_try_month
    financing_liability = installed_cs24 * financed_per_cs24_try_month
    equity_proxy = max(0, cumulative_cash) + installed_cs24 * (cs24_total_capex_try_month - financed_per_cs24_try_month)

    results.append({
        "Month": month_no,
        "Year": year,
        "Half": half,
        "Courier E-Scooter (Facelift) Units": courier_facelift_units,
        "Courier E-Scooter (New) Units": courier_new_units,
        "Commuter E-Scooter Units": commuter_units,
        "Monthly Production Units": monthly_production_units,
        "Units Sold": units_sold_month,
        "Units Added To Stock": units_added_to_stock,
        "Ending Stock Units": inventory_units,
        "Active E-Scooters": active_escooters,
        "USD/TRY (Month)": usd_try_month,
        "Inflation Rate (Month)": inflation_rate_month,
        "Energy Buy Price (Month)": energy_buy_price_month,
        "Energy Sell Price (Month)": energy_sell_price_month,
        "Monthly kWh Sold": monthly_kwh_sold,
        "Monthly Revenue": monthly_revenue,
        "Monthly Energy Cost": monthly_energy_cost,
        "Monthly Energy Gross Profit (NewCo)": monthly_energy_gp_newco,
        "Monthly Motor Gross Profit": monthly_motor_gross_profit,
        "Monthly CS24 Other Gross Profit": monthly_cs24_other_gp,
        "Monthly Total Gross Profit": monthly_total_gross_profit,
        "Required CS24": monthly_cs24_required,
        "New CS24": monthly_new_cs24,
        "CS24 CapEx Outflow": monthly_cs24_capex_outflow,
        "Raw Material Cost": raw_material_cost_month,
        "Total Personnel Cost": total_personnel_cost_month,
        "Total OPEX": total_opex_month,
        "Monthly EBITDA Proxy": monthly_ebitda_proxy,
        "Monthly Interest Cost": monthly_interest_cost,
        "Monthly Tax": monthly_tax,
        "Monthly Net Income": monthly_net_income,
        "Monthly Net Cash Flow": monthly_net_cash_flow,
        "Cumulative Cash Balance": cumulative_cash,
        "Raw Material Inventory": inventory_raw_material_value,
        "Finished Goods Inventory": inventory_units,
        "Installed CS24 Asset Value": installed_cs24_asset_value,
        "Financing Liability": financing_liability,
        "Equity Proxy": equity_proxy,
    })

monthly_df = pd.DataFrame(results)
personnel_detail_df = pd.DataFrame(personnel_records) if personnel_records else pd.DataFrame()

annual_df = monthly_df.groupby("Year", as_index=False).agg({
    "Monthly Production Units": "sum",
    "Units Sold": "sum",
    "Active E-Scooters": "last",
    "Monthly kWh Sold": "sum",
    "Monthly Revenue": "sum",
    "Monthly Energy Cost": "sum",
    "Monthly Energy Gross Profit (NewCo)": "sum",
    "Monthly Motor Gross Profit": "sum",
    "Monthly CS24 Other Gross Profit": "sum",
    "Monthly Total Gross Profit": "sum",
    "Total Personnel Cost": "sum",
    "Total OPEX": "sum",
    "Monthly EBITDA Proxy": "sum",
    "CS24 CapEx Outflow": "sum",
    "Raw Material Cost": "sum",
    "Monthly Net Cash Flow": "sum",
    "Cumulative Cash Balance": "last",
    "Required CS24": "last",
    "Installed CS24 Asset Value": "last",
    "Financing Liability": "last",
    "Equity Proxy": "last",
})

cash_flow_df = monthly_df[[
    "Month",
    "Monthly Revenue",
    "Monthly Energy Cost",
    "Monthly Energy Gross Profit (NewCo)",
    "Monthly Motor Gross Profit",
    "Monthly CS24 Other Gross Profit",
    "Monthly Total Gross Profit",
    "Total Personnel Cost",
    "Total OPEX",
    "Monthly EBITDA Proxy",
    "CS24 CapEx Outflow",
    "Raw Material Cost",
    "Monthly Interest Cost",
    "Monthly Tax",
    "Monthly Net Income",
    "Monthly Net Cash Flow",
    "Cumulative Cash Balance",
]].copy()

balance_sheet_df = monthly_df[[
    "Month",
    "Cumulative Cash Balance",
    "Raw Material Inventory",
    "Finished Goods Inventory",
    "Installed CS24 Asset Value",
    "Financing Liability",
    "Equity Proxy",
]].copy()


# =========================================================
# ASSUMPTIONS DISPLAY TABLE
# =========================================================

assumption_display_rows = []
for row in assumptions_layout:
    if row["kind"] == "section":
        assumption_display_rows.append({
            "Section": row["label"],
            "Parameter": "",
            "Value": "",
            "Type": "",
            "SortKey": row["row_no"],
            "Kind": "section"
        })
    else:
        label = row["label"]
        if label in calculated_map:
            val = calculated_map[label]
            typ = "Calculated"
            decimal_flag = row["decimal"]
        else:
            val = get_assumption(label, 0)
            typ = assumption_inputs[label]["type"].capitalize() if label in assumption_inputs else "Editable"
            decimal_flag = row["decimal"]

        if "%" in label:
            display_val = fmt_pct(val)
        elif "Price" in label or "Profit" in label:
            display_val = f"{safe_float(val, 0):,.1f}" if decimal_flag else fmt_try(val)
        elif "USD" in label:
            display_val = f"{safe_float(val, 0):,.1f}" if decimal_flag else fmt_num(val)
        else:
            display_val = f"{safe_float(val, 0):,.1f}" if decimal_flag else fmt_num(val)

        assumption_display_rows.append({
            "Section": "",
            "Parameter": label,
            "Value": display_val,
            "Type": typ,
            "SortKey": row["row_no"],
            "Kind": "line"
        })

assumptions_display_df = pd.DataFrame(assumption_display_rows).sort_values("SortKey")
assumptions_display_df = assumptions_display_df[["Section", "Parameter", "Value", "Type", "Kind"]]


# =========================================================
# DISPLAY FORMATTING
# =========================================================

dashboard_monthly_fmt = monthly_df.copy()
for c in dashboard_monthly_fmt.columns:
    dashboard_monthly_fmt[c] = dashboard_monthly_fmt[c].astype(object)

currency_cols = [
    "Monthly Revenue", "Monthly Energy Cost", "Monthly Energy Gross Profit (NewCo)",
    "Monthly Motor Gross Profit", "Monthly CS24 Other Gross Profit", "Monthly Total Gross Profit",
    "CS24 CapEx Outflow", "Raw Material Cost", "Total Personnel Cost", "Total OPEX",
    "Monthly EBITDA Proxy", "Monthly Interest Cost", "Monthly Tax", "Monthly Net Income",
    "Monthly Net Cash Flow", "Cumulative Cash Balance", "Raw Material Inventory",
    "Installed CS24 Asset Value", "Financing Liability", "Equity Proxy"
]
num_cols = [
    "Courier E-Scooter (Facelift) Units", "Courier E-Scooter (New) Units",
    "Commuter E-Scooter Units", "Monthly Production Units", "Units Sold", "Units Added To Stock",
    "Ending Stock Units", "Active E-Scooters", "USD/TRY (Month)", "Monthly kWh Sold", "Required CS24",
    "New CS24", "Finished Goods Inventory"
]
decimal_cols = ["Inflation Rate (Month)", "Energy Buy Price (Month)", "Energy Sell Price (Month)"]
plain_cols = ["Month", "Year"]

for c in plain_cols:
    if c in dashboard_monthly_fmt.columns:
        dashboard_monthly_fmt[c] = dashboard_monthly_fmt[c].map(fmt_plain_int)

for c in currency_cols:
    if c in dashboard_monthly_fmt.columns:
        dashboard_monthly_fmt[c] = dashboard_monthly_fmt[c].map(fmt_try)

for c in num_cols:
    if c in dashboard_monthly_fmt.columns:
        dashboard_monthly_fmt[c] = dashboard_monthly_fmt[c].map(fmt_num)

for c in decimal_cols:
    if c in dashboard_monthly_fmt.columns:
        dashboard_monthly_fmt[c] = dashboard_monthly_fmt[c].map(lambda x: f"{safe_float(x, 0):,.1f}")

annual_fmt = annual_df.copy()
for c in annual_fmt.columns:
    annual_fmt[c] = annual_fmt[c].astype(object)

for c in annual_fmt.columns:
    if c == "Year":
        annual_fmt[c] = annual_fmt[c].map(fmt_plain_int)
    elif c in ["Monthly Production Units", "Units Sold", "Active E-Scooters", "Monthly kWh Sold", "Required CS24"]:
        annual_fmt[c] = annual_fmt[c].map(fmt_num)
    else:
        annual_fmt[c] = annual_fmt[c].map(fmt_try)

cash_flow_fmt = cash_flow_df.copy()
for c in cash_flow_fmt.columns:
    cash_flow_fmt[c] = cash_flow_fmt[c].astype(object)

for c in cash_flow_fmt.columns:
    if c == "Month":
        cash_flow_fmt[c] = cash_flow_fmt[c].map(fmt_plain_int)
    else:
        cash_flow_fmt[c] = cash_flow_fmt[c].map(fmt_try)

balance_fmt = balance_sheet_df.copy()
for c in balance_fmt.columns:
    balance_fmt[c] = balance_fmt[c].astype(object)

for c in balance_fmt.columns:
    if c == "Month":
        balance_fmt[c] = balance_fmt[c].map(fmt_plain_int)
    elif c == "Finished Goods Inventory":
        balance_fmt[c] = balance_fmt[c].map(fmt_num)
    else:
        balance_fmt[c] = balance_fmt[c].map(fmt_try)

personnel_fmt = personnel_detail_df.copy()
if not personnel_fmt.empty:
    for c in personnel_fmt.columns:
        personnel_fmt[c] = personnel_fmt[c].astype(object)

    if "Month" in personnel_fmt.columns:
        personnel_fmt["Month"] = personnel_fmt["Month"].map(fmt_plain_int)
    if "Headcount" in personnel_fmt.columns:
        personnel_fmt["Headcount"] = personnel_fmt["Headcount"].map(fmt_num)
    if "Monthly Salary (Actual)" in personnel_fmt.columns:
        personnel_fmt["Monthly Salary (Actual)"] = personnel_fmt["Monthly Salary (Actual)"].map(fmt_try)
    if "Monthly Personnel Cost" in personnel_fmt.columns:
        personnel_fmt["Monthly Personnel Cost"] = personnel_fmt["Monthly Personnel Cost"].map(fmt_try)


# =========================================================
# TABS
# =========================================================

tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
    "Assumptions",
    "Investor Dashboard",
    "Monthly Financials",
    "Cash Flow Statement",
    "Balance Sheet",
    "Personnel",
    "Excel Preview",
])

with tab1:
    st.subheader("Assumptions")

    for _, r in assumptions_display_df.iterrows():
        if r["Kind"] == "section":
            st.markdown(f"**{r['Section']}**")
        else:
            c1, c2, c3 = st.columns([5, 2, 2])
            c1.write(r["Parameter"])
            c2.write(r["Value"])
            c3.write(r["Type"])

with tab2:
    st.subheader("Investor Dashboard")

    last_row = monthly_df.iloc[-1]

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("Final Active E-Scooters", fmt_num(last_row["Active E-Scooters"]))
    c2.metric("Final Monthly kWh Sold", fmt_num(last_row["Monthly kWh Sold"]))
    c3.metric("Final Monthly Revenue", fmt_try(last_row["Monthly Revenue"]))
    c4.metric("Final Monthly EBITDA Proxy", fmt_try(last_row["Monthly EBITDA Proxy"]))

    c5, c6, c7, c8 = st.columns(4)
    c5.metric("Final Required CS24", fmt_num(last_row["Required CS24"]))
    c6.metric("Final Cumulative Cash", fmt_try(last_row["Cumulative Cash Balance"]))
    c7.metric("Raw Material Lag", fmt_num(raw_material_required_before_months))
    c8.metric("Stock-In Rate", fmt_pct(stock_in_rate_pct))

    st.markdown("### Annual Summary")
    st.dataframe(annual_fmt, use_container_width=True, hide_index=True)

    st.markdown("### Monthly Trend")
    st.line_chart(monthly_df.set_index("Month")[["Monthly Revenue", "Monthly EBITDA Proxy", "Monthly Net Cash Flow"]])

with tab3:
    st.subheader("Monthly Financials")
    st.dataframe(dashboard_monthly_fmt, use_container_width=True, hide_index=True)

with tab4:
    st.subheader("Cash Flow Statement")
    st.dataframe(cash_flow_fmt, use_container_width=True, hide_index=True)
    st.line_chart(cash_flow_df.set_index("Month")[["Monthly Net Cash Flow", "Cumulative Cash Balance"]])

with tab5:
    st.subheader("Balance Sheet")
    st.dataframe(balance_fmt, use_container_width=True, hide_index=True)
    st.line_chart(balance_sheet_df.set_index("Month")[[
        "Cumulative Cash Balance", "Raw Material Inventory",
        "Installed CS24 Asset Value", "Financing Liability", "Equity Proxy"
    ]])

with tab6:
    st.subheader("Personnel")
    if personnel_fmt.empty:
        st.info("No personnel data available.")
    else:
        st.dataframe(personnel_fmt, use_container_width=True, hide_index=True)

with tab7:
    st.subheader("Excel Preview")
    if monthly_inputs_df is not None:
        st.markdown("### Monthly Inputs")
        st.dataframe(monthly_inputs_df, use_container_width=True)
    if macro_df is not None:
        st.markdown("### Macro Assumptions")
        st.dataframe(macro_df, use_container_width=True)
    if opex_df is not None:
        st.markdown("### OPEX")
        st.dataframe(opex_df, use_container_width=True)
    if personnel_base_df is not None:
        st.markdown("### Personnel Base")
        st.dataframe(personnel_base_df, use_container_width=True)
    if personnel_rules_df is not None:
        st.markdown("### Personnel Monthly Rules")
        st.dataframe(personnel_rules_df, use_container_width=True)
    if bom_df is not None:
        st.markdown("### BoM List")
        st.dataframe(bom_df, use_container_width=True)
    if product_capex_df is not None:
        st.markdown("### Product CapEx")
        st.dataframe(product_capex_df, use_container_width=True)
